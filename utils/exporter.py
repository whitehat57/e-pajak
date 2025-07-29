import csv
import json
from datetime import datetime
from typing import List, Dict
from models.employee import Employee
from models.transaction import Transaction
from models.tax import TaxRecord
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class ReportExporter:
    def __init__(self):
        # Buat direktori export jika belum ada
        self.export_dir = "exports"
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir)
    
    def export_employees_to_csv(self) -> str:
        """Ekspor data pegawai ke CSV"""
        employees = Employee.get_all()
        
        if not employees:
            raise Exception("Tidak ada data pegawai untuk diekspor")
        
        # Buat nama file dengan timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.export_dir}/pegawai_{timestamp}.csv"
        
        # Header CSV
        fieldnames = ['ID', 'Nama', 'Status', 'Gaji_Bulanan', 'Tunjangan', 'NPWP', 'Tanggal_Dibuat']
        
        # Ekspor ke CSV
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for emp in employees:
                writer.writerow({
                    'ID': emp.id,
                    'Nama': emp.name,
                    'Status': emp.status,
                    'Gaji_Bulanan': emp.monthly_salary,
                    'Tunjangan': emp.allowances,
                    'NPWP': emp.npwp or '',
                    'Tanggal_Dibuat': emp.created_at
                })
        
        return filename
    
    def export_employees_to_excel(self) -> str:
        """Ekspor data pegawai ke Excel"""
        employees = Employee.get_all()
        
        if not employees:
            raise Exception("Tidak ada data pegawai untuk diekspor")
        
        # Buat nama file dengan timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.export_dir}/pegawai_{timestamp}.xlsx"
        
        # Buat workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Pegawai"
        
        # Header
        headers = ['ID', 'Nama', 'Status', 'Gaji Bulanan', 'Tunjangan', 'NPWP', 'Tanggal Dibuat']
        ws.append(headers)
        
        # Style untuk header
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for emp in employees:
            ws.append([
                emp.id,
                emp.name,
                emp.status.title(),
                emp.monthly_salary,
                emp.allowances,
                emp.npwp or '-',
                emp.created_at
            ])
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        return filename
    
    def export_transactions_to_csv(self) -> str:
        """Ekspor data transaksi ke CSV"""
        transactions = Transaction.get_all()
        
        if not transactions:
            raise Exception("Tidak ada data transaksi untuk diekspor")
        
        # Buat nama file dengan timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.export_dir}/transaksi_{timestamp}.csv"
        
        # Header CSV
        fieldnames = ['ID', 'Jenis', 'Deskripsi', 'Jumlah', 'PPN', 'Total', 'Tanggal', 'No_Faktur', 'Tanggal_Dibuat']
        
        # Ekspor ke CSV
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for trans in transactions:
                total_amount = trans.amount + trans.ppn_amount
                writer.writerow({
                    'ID': trans.id,
                    'Jenis': trans.type,
                    'Deskripsi': trans.description,
                    'Jumlah': trans.amount,
                    'PPN': trans.ppn_amount,
                    'Total': total_amount,
                    'Tanggal': trans.transaction_date,
                    'No_Faktur': trans.invoice_number or '',
                    'Tanggal_Dibuat': trans.created_at
                })
        
        return filename
    
    def export_transactions_to_excel(self) -> str:
        """Ekspor data transaksi ke Excel"""
        transactions = Transaction.get_all()
        
        if not transactions:
            raise Exception("Tidak ada data transaksi untuk diekspor")
        
        # Buat nama file dengan timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.export_dir}/transaksi_{timestamp}.xlsx"
        
        # Buat workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Transaksi"
        
        # Header
        headers = ['ID', 'Jenis', 'Deskripsi', 'Jumlah', 'PPN', 'Total', 'Tanggal', 'No. Faktur', 'Tanggal Dibuat']
        ws.append(headers)
        
        # Style untuk header
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for trans in transactions:
            total_amount = trans.amount + trans.ppn_amount
            ws.append([
                trans.id,
                trans.type.title(),
                trans.description,
                trans.amount,
                trans.ppn_amount,
                total_amount,
                trans.transaction_date,
                trans.invoice_number or '-',
                trans.created_at
            ])
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        return filename
    
    def export_tax_records_to_csv(self) -> str:
        """Ekspor data catatan pajak ke CSV"""
        tax_records = TaxRecord.get_all()
        
        if not tax_records:
            raise Exception("Tidak ada data catatan pajak untuk diekspor")
        
        # Buat nama file dengan timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.export_dir}/catatan_pajak_{timestamp}.csv"
        
        # Header CSV
        fieldnames = ['ID', 'Nama_Pegawai', 'Periode', 'Jenis_Pajak', 'Penghasilan_Bruto', 'PKP', 'Pajak_Terutang', 'Deskripsi', 'Tanggal_Dibuat']
        
        # Ekspor ke CSV
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for record in tax_records:
                # Dapatkan nama pegawai jika ada
                employee_name = "-"
                if record.employee_id:
                    employee = Employee.get_by_id(record.employee_id)
                    if employee:
                        employee_name = employee.name
                
                writer.writerow({
                    'ID': record.id,
                    'Nama_Pegawai': employee_name,
                    'Periode': record.period,
                    'Jenis_Pajak': record.tax_type,
                    'Penghasilan_Bruto': record.gross_income,
                    'PKP': record.taxable_income,
                    'Pajak_Terutang': record.tax_amount,
                    'Deskripsi': record.description,
                    'Tanggal_Dibuat': record.created_at
                })
        
        return filename
    
    def export_tax_records_to_excel(self) -> str:
        """Ekspor data catatan pajak ke Excel"""
        tax_records = TaxRecord.get_all()
        
        if not tax_records:
            raise Exception("Tidak ada data catatan pajak untuk diekspor")
        
        # Buat nama file dengan timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.export_dir}/catatan_pajak_{timestamp}.xlsx"
        
        # Buat workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Catatan Pajak"
        
        # Header
        headers = ['ID', 'Nama Pegawai', 'Periode', 'Jenis Pajak', 'Penghasilan Bruto', 'PKP', 'Pajak Terutang', 'Deskripsi', 'Tanggal Dibuat']
        ws.append(headers)
        
        # Style untuk header
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for record in tax_records:
            # Dapatkan nama pegawai jika ada
            employee_name = "-"
            if record.employee_id:
                employee = Employee.get_by_id(record.employee_id)
                if employee:
                    employee_name = employee.name
            
            ws.append([
                record.id,
                employee_name,
                record.period,
                record.tax_type.upper(),
                record.gross_income,
                record.taxable_income,
                record.tax_amount,
                record.description,
                record.created_at
            ])
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        return filename
    
    def export_spt_summary_to_excel(self, spt_data: Dict) -> str:
        """Ekspor ringkasan SPT ke Excel dengan multiple sheet"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.export_dir}/spt_tahunan_{spt_data['year']}_{timestamp}.xlsx"
        
        # Buat workbook
        wb = Workbook()
        
        # Sheet 1: Ringkasan SPT
        ws1 = wb.active
        ws1.title = "Ringkasan SPT"
        
        # Header
        ws1.append(['Komponen', 'Nilai'])
        
        # Style untuk header
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws1.cell(row=1, column=1).font = header_font
        ws1.cell(row=1, column=1).fill = header_fill
        ws1.cell(row=1, column=2).font = header_font
        ws1.cell(row=1, column=2).fill = header_fill
        
        # Data ringkasan
        summary_data = [
            ['Tahun Pelaporan', spt_data['year']],
            ['Total Penjualan', spt_data['income_summary']['total_sales']],
            ['PPN Keluaran', spt_data['income_summary']['total_sales_ppn']],
            ['Total Pembelian', spt_data['income_summary']['total_purchases']],
            ['PPN Masukan', spt_data['income_summary']['total_purchases_ppn']],
            ['Penghasilan Bruto', spt_data['income_summary']['gross_income']],
            ['Biaya Usaha', spt_data['income_summary']['business_expenses']],
            ['Penghasilan Neto', spt_data['income_summary']['net_income']],
            ['Jumlah Pegawai', spt_data['employee_summary']['total_employees']],
            ['Total Gaji Pokok', spt_data['employee_summary']['total_gross_salary']],
            ['Total Tunjangan', spt_data['employee_summary']['total_allowances']],
            ['PPh 21 Dipotong', spt_data['employee_summary']['total_pph21_withheld']],
            ['PPN Terutang', spt_data['ppn_summary']['ppn_payable']],
            ['Penghasilan Kena Pajak', spt_data['tax_calculation']['taxable_income']],
            ['PPh Badan Terutang', spt_data['tax_calculation']['corporate_tax_payable']],
            ['PPh 21 Telah Dibayar', spt_data['tax_calculation']['tax_paid']],
            ['PPh Badan Terutang Bersih', spt_data['tax_calculation']['net_tax_payable']]
        ]
        
        for row_data in summary_data:
            ws1.append(row_data)
        
        # Auto-adjust column width
        for column in ws1.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Detail Pegawai
        ws2 = wb.create_sheet("Detail Pegawai")
        
        # Header
        employee_headers = ['ID', 'Nama', 'Status', 'Gaji Bulanan', 'Tunjangan Bulanan', 'Gaji Tahunan', 'NPWP']
        ws2.append(employee_headers)
        
        # Style untuk header
        for col in range(1, len(employee_headers) + 1):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data pegawai
        employees = Employee.get_all()
        for emp in employees:
            ws2.append([
                emp.id,
                emp.name,
                emp.status.title(),
                emp.monthly_salary,
                emp.allowances,
                (emp.monthly_salary + emp.allowances) * 12,
                emp.npwp or '-'
            ])
        
        # Auto-adjust column width
        for column in ws2.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 3: Detail Transaksi
        ws3 = wb.create_sheet("Detail Transaksi")
        
        # Header
        transaction_headers = ['Tanggal', 'Jenis', 'Deskripsi', 'Jumlah', 'PPN', 'Total', 'No. Faktur']
        ws3.append(transaction_headers)
        
        # Style untuk header
        for col in range(1, len(transaction_headers) + 1):
            cell = ws3.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data transaksi
        transactions = Transaction.get_all()
        for trans in transactions:
            ws3.append([
                trans.transaction_date,
                trans.type.title(),
                trans.description,
                trans.amount,
                trans.ppn_amount,
                trans.amount + trans.ppn_amount,
                trans.invoice_number or '-'
            ])
        
        # Auto-adjust column width
        for column in ws3.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws3.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        return filename
    
    def get_export_history(self) -> List[Dict]:
        """Dapatkan riwayat ekspor file"""
        history = []
        
        if os.path.exists(self.export_dir):
            for filename in os.listdir(self.export_dir):
                if filename.endswith(('.csv', '.xlsx')):
                    file_path = os.path.join(self.export_dir, filename)
                    file_stat = os.stat(file_path)
                    history.append({
                        'filename': filename,
                        'filepath': file_path,
                        'size': file_stat.st_size,
                        'modified': datetime.fromtimestamp(file_stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                    })
        
        # Urutkan berdasarkan tanggal modifikasi (terbaru dulu)
        history.sort(key=lambda x: x['modified'], reverse=True)
        return history
    
    def delete_export_file(self, filename: str) -> bool:
        """Hapus file ekspor"""
        file_path = os.path.join(self.export_dir, filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            return True
        return False